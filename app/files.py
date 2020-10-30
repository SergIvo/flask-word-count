#import csv
import json

#from io import BytesIO, StringIO
from string import ascii_letters
from random import choice
from os import listdir, remove
from os.path import getmtime, getsize
from sys import getsizeof
from time import time
from tempfile import NamedTemporaryFile

import openpyxl

def check_expiration(dir, maxtime = None):
    if not maxtime:
        maxtime = 1800
    files = listdir(dir)
    times = {file: (time() - getmtime(f'{dir}/{file}')) for file in files}
    for file in times:
        if times[file] > maxtime:
            remove(f'{dir}/{file}')
            files.remove(file)
    
def check_size(dir, object, maxsize = None):
    if not maxsize:
        maxsize = 20971520
    files = listdir(dir)
    total_size = sum((getsize(f'{dir}/{file}') for file in files))
    obj_size = getsizeof(json.dumps(object))
    if obj_size >= maxsize:
        return False
    if obj_size + total_size > maxsize:
        sorted(files, key = lambda file: getmtime(f'{dir}/{file}'))
        while obj_size + total_size > maxsize:
            remove(f'{dir}/{files[0]}')
            files.pop(0)
            total_size = sum((getsize(f'{dir}/{file}') for file in files))
    return total_size, obj_size
    

def temp_name(extension):
    combination = ''.join(choice(ascii_letters) for i in range(8))
    return f'temp_{combination}.{extension}'

def save_json(data, name):
    with open(name, 'w') as file:
        json.dump(data, file)
        
def from_json(name):
    with open(name, 'r') as file:
        data  =json.load(file)
        return data

def save_temp_xlsx(data):
    wb = openpyxl.Workbook()
    keylist = [key for key in data.keys()]
    for key in data:
        if keylist.index(key) == 0:
            sheet = wb.active
            sheet.title = key
        else:
            sheet = wb.create_sheet(title = key)
        if isinstance(data[key], list):
            sheet.append(list(data[key][0].keys()))
            for row in data[key]:
                sheet.append(list(row.values()))
        elif isinstance(data[key], dict):
            current_row = 1
            for title_key in data[key]:
                sheet[f'A{current_row}'] = title_key
                sheet.merge_cells(f'A{current_row}:B{current_row}')
                current_row += 1
                for param in data[key][title_key]:
                    sheet[f'A{current_row}'] = param
                    sheet[f'B{current_row}'] = data[key][title_key][param]
                    current_row += 1
                current_row += 1
    temp = NamedTemporaryFile()
    wb.save(temp.name)
    return temp

#def produce_csv(data):
#    csv_file = StringIO()
#    writer = csv.DictWriter(csv_file, fieldnames = data[0].keys())
#    for row in data:
#        writer.writerow(row)
#    file_bytes = BytesIO()
#    file_bytes.write(csv_file.getvalue().encode('utf-8'))
#    file_bytes.seek(0)
#    csv_file.close()
#    return file_bytes

#https://realpython.com/openpyxl-excel-spreadsheets-python/
#https://develop.zendesk.com/hc/en-us/articles/360001074408-Writing-large-data-sets-to-Excel-with-Python-and-pandas
#https://stackoverflow.com/questions/8469665/saving-openpyxl-file-via-text-and-filestream/55144731
#https://habr.com/ru/company/otus/blog/331998/

#https://stackoverflow.com/questions/31391344/using-tempfile-to-create-pdf-xls-documents-in-flask
#https://stackoverflow.com/questions/13344538/how-to-clean-up-temporary-file-used-with-send-file
#https://stackoverflow.com/questions/14614756/how-can-i-generate-file-on-the-fly-and-delete-it-after-download
